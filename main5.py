"""
SpeakSense — Professional Meeting Speech Analyser
Template: Timeline · Speakers · Speaker-wise Summary · Key Points ·
          Action Items · Overall Summary · Insights
Requirements: customtkinter, openai-whisper, torch, sounddevice,
              numpy, openpyxl, ollama, librosa, scikit-learn,
              soundfile, pydub (optional for non-wav)
"""

import re, uuid, datetime, threading
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import tkinter as tk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
import customtkinter as ctk

# ─── Theme: Light / White ─────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ─── Palette ──────────────────────────────────────────────────────────────────
BG        = "#F7F9FC"
SURFACE   = "#FFFFFF"
CARD      = "#FFFFFF"
RAISED    = "#F0F4FA"
BORDER    = "#DDE3EF"
TEXT      = "#1A2035"
MUTED     = "#6B7A99"
ACCENT    = "#3B5BDB"
RED       = "#E03131"
GREEN     = "#2F9E44"
AMBER     = "#E67700"
TEAL      = "#0C8599"
PURPLE    = "#7048E8"
LIGHT_RED = "#FFF5F5"
LIGHT_GRN = "#EBFBEE"
LIGHT_AMB = "#FFF9DB"
LIGHT_TEL = "#E6FCF5"
LIGHT_PUR = "#F3F0FF"
LIGHT_BLU = "#EEF2FF"

TOPBAR_BG  = "#1E2D6B"
TOPBAR_TXT = "#FFFFFF"

# ─── Speaker Colours (6 distinct) ─────────────────────────────────────────────
SPK_FG = ["#1971C2","#2F9E44","#E67700","#7048E8","#0C8599","#C2255C"]
SPK_BG = ["#E7F5FF","#EBFBEE","#FFF3BF","#F3F0FF","#E6FCF5","#FFE3EC"]
SPK_BD = ["#74C0FC","#8CE99A","#FFE066","#B197FC","#63E6BE","#F783AC"]

def spk_fg(i):    return SPK_FG[i % len(SPK_FG)]
def spk_bg(i):    return SPK_BG[i % len(SPK_BG)]
def spk_bd(i):    return SPK_BD[i % len(SPK_BD)]
def spk_label(i): return f"Speaker {i + 1}"

def fmt_time(s: float) -> str:
    m = int(s) // 60; sec = int(s) % 60
    return f"{m:02d}:{sec:02d}"


# ─── Improved Diarizer ────────────────────────────────────────────────────────
class SimpleDiarizer:
    """
    Speaker diarizer using rich acoustic features:
      • MFCC (20 coeffs) + delta + delta-delta
      • Spectral centroid, rolloff, bandwidth
      • Zero-crossing rate
      • Pitch (fundamental frequency via piptrack)
    Clustering: Agglomerative (Ward linkage) → smoother than KMeans.
    Post-processing: label smoothing removes isolated 1-segment flips.
    """

    def __init__(self, n_speakers: int = 2):
        self.n_speakers = n_speakers

    # ── Public API ────────────────────────────────────────────────────────
    def diarize(self, wav_path: str, segments: List[Dict]) -> List[Dict]:
        if not segments:
            return segments
        try:
            import librosa
            from sklearn.cluster import AgglomerativeClustering, KMeans
            from sklearn.preprocessing import StandardScaler, normalize
            from sklearn.decomposition import PCA

            audio, sr = librosa.load(wav_path, sr=16000, mono=True)
            feats = self._extract_features(audio, sr, segments)

            X = StandardScaler().fit_transform(np.array(feats))

            # Reduce dimensionality only when we have enough samples
            if X.shape[0] >= 6 and X.shape[1] > 20:
                n_comp = min(20, X.shape[0] - 1, X.shape[1])
                X = PCA(n_components=n_comp, random_state=42).fit_transform(X)

            n_clusters = min(self.n_speakers, len(segments))
            labels = self._cluster(X, n_clusters)

            for seg, lbl in zip(segments, labels):
                seg["speaker"] = int(lbl)

            # Post-process: smooth out isolated noise flips
            self._smooth_labels(segments, window=3)
            # Re-map label indices to be contiguous (0, 1, 2…)
            self._remap_labels(segments)

        except Exception as exc:
            print(f"[Diarizer] Falling back to gap-based: {exc}")
            self._gap_fallback(segments)

        return segments

    # ── Feature Extraction ────────────────────────────────────────────────
    def _extract_features(self, audio: np.ndarray, sr: int,
                          segments: List[Dict]) -> np.ndarray:
        """Return (n_segments, n_features) array of acoustic features."""
        import librosa
        all_feats = []
        FEAT_DIM = 20 * 4 + 4   # 84 dims — consistent for every segment

        for seg in segments:
            s = int(seg["start"] * sr)
            e = int(seg["end"] * sr)
            chunk = audio[s:e]

            if len(chunk) < 1024:
                # Too short — use silence-like zero vector
                all_feats.append(np.zeros(FEAT_DIM))
                continue

            try:
                # ── MFCCs (20) + statistics ────────────────────────────
                mfcc = librosa.feature.mfcc(y=chunk, sr=sr, n_mfcc=20)
                mfcc_mean  = mfcc.mean(axis=1)           # (20,)
                mfcc_std   = mfcc.std(axis=1)            # (20,)
                delta1     = librosa.feature.delta(mfcc).mean(axis=1)   # (20,)
                delta2     = librosa.feature.delta(mfcc, order=2).mean(axis=1)  # (20,)

                # ── Spectral features (4 scalars) ─────────────────────
                spec_cent = float(librosa.feature.spectral_centroid(
                    y=chunk, sr=sr).mean())
                spec_bw   = float(librosa.feature.spectral_bandwidth(
                    y=chunk, sr=sr).mean())
                spec_roll = float(librosa.feature.spectral_rolloff(
                    y=chunk, sr=sr).mean())
                zcr       = float(librosa.feature.zero_crossing_rate(chunk).mean())

                # ── Pitch (fundamental freq) ──────────────────────────
                pitches, mags = librosa.piptrack(y=chunk, sr=sr, fmin=60, fmax=600)
                voiced = pitches[mags > mags.mean()]
                pitch_mean = float(voiced.mean()) if voiced.size > 0 else 0.0

                feat = np.concatenate([
                    mfcc_mean, mfcc_std, delta1, delta2,
                    [spec_cent, spec_bw, spec_roll, zcr]   # 4 scalars → total 84
                ])
                # Replace pitch_mean into the last slot for consistency
                feat[-1] = pitch_mean

            except Exception:
                feat = np.zeros(FEAT_DIM)

            all_feats.append(feat)

        return np.array(all_feats, dtype=np.float32)

    # ── Clustering ────────────────────────────────────────────────────────
    def _cluster(self, X: np.ndarray, n_clusters: int) -> np.ndarray:
        from sklearn.cluster import AgglomerativeClustering, KMeans

        if n_clusters == 1:
            return np.zeros(len(X), dtype=int)

        # Ward / Agglomerative is generally better than KMeans for speaker diarization
        try:
            labels = AgglomerativeClustering(
                n_clusters=n_clusters,
                linkage="ward"
            ).fit_predict(X)
        except Exception:
            # Fallback to KMeans if agglomerative fails
            labels = KMeans(
                n_clusters=n_clusters,
                random_state=42,
                n_init=15,
                max_iter=500
            ).fit_predict(X)

        return labels

    # ── Label Smoothing ───────────────────────────────────────────────────
    def _smooth_labels(self, segments: List[Dict], window: int = 3):
        """
        Three-pass smoothing:
        1. Single-segment islands: if surrounded by same speaker, absorb.
        2. Short runs (< min_run): merge into neighbours.
        3. Enforce a minimum pause between speaker switches.
        """
        n = len(segments)
        if n < 3:
            return

        labels = [s["speaker"] for s in segments]

        # Pass 1 — single-island removal (3 iterations)
        for _ in range(3):
            for i in range(1, n - 1):
                if labels[i - 1] == labels[i + 1] and labels[i] != labels[i - 1]:
                    labels[i] = labels[i - 1]

        # Pass 2 — short run (≤ 2 segments) surrounded by the same speaker
        i = 0
        while i < n:
            j = i
            while j < n and labels[j] == labels[i]:
                j += 1
            run_len = j - i
            if run_len <= 2 and i > 0 and j < n:
                if labels[i - 1] == labels[j]:
                    for k in range(i, j):
                        labels[k] = labels[i - 1]
            i = j

        # Pass 3 — time-gap guard: if a segment starts < 0.4 s after the previous
        # ends, keep the same speaker (avoids splitting a breath pause)
        for i in range(1, n):
            gap = segments[i]["start"] - segments[i - 1]["end"]
            if gap < 0.4 and labels[i] != labels[i - 1]:
                labels[i] = labels[i - 1]

        for seg, lbl in zip(segments, labels):
            seg["speaker"] = lbl

    # ── Label Remapping ───────────────────────────────────────────────────
    def _remap_labels(self, segments: List[Dict]):
        """Ensure speaker indices are 0-based integers in order of first appearance."""
        mapping: Dict[int, int] = {}
        counter = 0
        for seg in segments:
            old = seg["speaker"]
            if old not in mapping:
                mapping[old] = counter
                counter += 1
            seg["speaker"] = mapping[old]

    # ── Gap-based Fallback ────────────────────────────────────────────────
    def _gap_fallback(self, segments: List[Dict]):
        gap, spk, prev = 2.0, 0, 0.0
        for seg in segments:
            if seg["start"] - prev > gap:
                spk = (spk + 1) % max(1, self.n_speakers)
            seg["speaker"] = spk
            prev = seg["end"]


# ─── Audio Processor ──────────────────────────────────────────────────────────
class AudioProcessor:
    def __init__(self, model_size: str = "base"):
        import whisper
        self.model = whisper.load_model(model_size)
        self.model_size = model_size

    def process(self, wav_path: str, n_speakers: int = 2) -> Dict:
        # Always transcribe without forcing a language — let Whisper auto-detect
        r = self.model.transcribe(
            str(wav_path),
            word_timestamps=True,
            verbose=False,
            fp16=False,
            task="transcribe"
        )
        segs = [{
            "id": s["id"],
            "start": round(s["start"], 2),
            "end":   round(s["end"],   2),
            "text":  s["text"].strip(),
            "speaker": 0,
        } for s in r["segments"] if s["text"].strip()]

        segs = SimpleDiarizer(n_speakers).diarize(wav_path, segs)
        spk_data = self._build_speakers(segs)
        return {
            "segments":     segs,
            "speaker_data": spk_data,
            "full_text":    " ".join(s["text"] for s in segs),
            "duration":     segs[-1]["end"] if segs else 0.0,
        }

    def _build_speakers(self, segs: List[Dict]) -> Dict[int, Dict]:
        data: Dict[int, Dict] = {}
        for s in segs:
            idx = s["speaker"]
            if idx not in data:
                data[idx] = {
                    "text": "", "segments": [], "duration": 0.0,
                    "count": 0, "first": s["start"], "last": s["end"]
                }
            d = data[idx]
            d["text"]    += (" " + s["text"]).lstrip()
            d["segments"].append({"start": s["start"], "end": s["end"], "text": s["text"]})
            d["duration"] += round(s["end"] - s["start"], 2)
            d["count"]    += 1
            d["last"]      = max(d["last"], s["end"])
        for d in data.values():
            d["text"] = d["text"].strip()
        return dict(sorted(data.items()))


# ─── Summariser ───────────────────────────────────────────────────────────────
class Summarizer:
    STOP = {
        "the","a","an","in","on","at","to","for","of","and","or","but","is","are",
        "was","were","be","been","it","this","that","with","i","we","you","he",
        "she","they","so","as","by","from","not","have","has","had","do","did",
        "will","would","could","should","may","might","just","also","then","than",
    }

    def _ollama(self, text: str, model: str, hint: str, max_w: int = 200) -> str:
        import ollama
        prompt = (
            f"You are a professional meeting analyst.\n{hint}\n"
            f"Limit response to {max_w} words.\n\nTEXT:\n{text[:5000]}\n\nRESPONSE:"
        )
        r = ollama.chat(model=model, messages=[{"role": "user", "content": prompt}])
        return r["message"]["content"].strip()

    def _extractive(self, text: str, n: int = 3) -> str:
        sents = [s.strip() for s in re.split(r"(?<=[.!?])\s+", text) if len(s.split()) >= 4]
        if not sents: return text[:400]
        freq = Counter(re.findall(r"\b\w+\b", text.lower()))
        scored = []
        for s in sents:
            words = [w for w in re.findall(r"\b\w+\b", s.lower()) if w not in self.STOP]
            scored.append((sum(freq.get(w, 0) for w in words) / (len(words) + 1), s))
        scored.sort(reverse=True)
        return " ".join(x[1] for x in sorted(scored[:n], key=lambda x: sents.index(x[1])))

    # ── Timeline ──────────────────────────────────────────────────────────
    def build_timeline(self, segments: List[Dict], duration: float) -> List[Dict]:
        if not segments: return []
        bucket_size = max(60.0, duration / 6)
        n_buckets   = max(1, int(np.ceil(duration / bucket_size)))
        buckets: List[Dict] = []
        for b in range(n_buckets):
            t0 = b * bucket_size; t1 = t0 + bucket_size
            chunk = " ".join(s["text"] for s in segments if t0 <= s["start"] < t1)
            if not chunk.strip(): continue
            buckets.append({"start": t0, "end": min(t1, duration), "text": chunk})
        return buckets

    def timeline_topics(self, buckets: List[Dict], model: str) -> List[str]:
        topics = []
        for b in buckets:
            try:
                t = self._ollama(b["text"], model,
                                 "Give a 5-8 word topic label for this audio segment. "
                                 "No preamble, just the label.", 15)
            except Exception:
                sents = b["text"].split(".")
                t = sents[0][:60].strip() if sents else "Discussion"
            topics.append(t)
        return topics

    # ── Speaker analysis ──────────────────────────────────────────────────
    def speaker_summary(self, text: str, idx: int, model: str) -> Dict:
        if len(text.split()) < 8:
            return {"discussed": text, "contribution": text, "tone": "Neutral"}
        try:
            raw = self._ollama(
                text, model,
                f"Analyse what {spk_label(idx)} said and respond in this EXACT format:\n"
                "DISCUSSED: <main topic in one sentence>\n"
                "CONTRIBUTION: <key contribution in one sentence>\n"
                "TONE: <one word: Formal/Informal/Analytical/Collaborative/Critical>",
                100)
            result = {"discussed": "", "contribution": "", "tone": "Neutral"}
            for line in raw.splitlines():
                if line.upper().startswith("DISCUSSED:"):
                    result["discussed"] = line.split(":", 1)[1].strip()
                elif line.upper().startswith("CONTRIBUTION:"):
                    result["contribution"] = line.split(":", 1)[1].strip()
                elif line.upper().startswith("TONE:"):
                    result["tone"] = line.split(":", 1)[1].strip()
            return result
        except Exception:
            return {"discussed": self._extractive(text, 1),
                    "contribution": self._extractive(text, 1), "tone": "Neutral"}

    # ── Key points ────────────────────────────────────────────────────────
    def key_points(self, full_text: str, model: str) -> List[str]:
        try:
            raw = self._ollama(full_text, model,
                               "Extract exactly 4-6 key points from this meeting. "
                               "Each on a new line starting with '•'. No preamble.", 150)
            pts = [l.lstrip("•-").strip() for l in raw.splitlines()
                   if l.strip() and l.strip()[0] in "•-"]
            return pts[:6] if pts else [self._extractive(full_text, 1)]
        except Exception:
            sents = [s.strip() for s in re.split(r"(?<=[.!?])\s+", full_text)
                     if len(s.split()) >= 6][:5]
            return sents

    # ── Action items ──────────────────────────────────────────────────────
    def action_items(self, full_text: str, model: str) -> List[str]:
        try:
            raw = self._ollama(full_text, model,
                               "List all action items, tasks, or next steps mentioned. "
                               "Each on a new line starting with '•'. "
                               "If none found, write '• No explicit action items mentioned.'",
                               150)
            items = [l.lstrip("•-").strip() for l in raw.splitlines()
                     if l.strip() and l.strip()[0] in "•-"]
            return items[:8] if items else ["No explicit action items mentioned."]
        except Exception:
            return ["Could not extract action items (Ollama not running)."]

    # ── Overall summary ───────────────────────────────────────────────────
    def overall_summary(self, full_text: str,
                        spk_summaries: Dict[int, Dict], model: str) -> str:
        combined = "\n".join(
            f"[{spk_label(i)}] discussed: {s.get('discussed','')} "
            f"| contributed: {s.get('contribution','')}"
            for i, s in spk_summaries.items())
        try:
            return self._ollama(
                combined, model,
                "Write one coherent paragraph (3-5 sentences) summarising the entire meeting. "
                "Cover: main topic, key discussion points, decisions made, and overall direction. "
                'Start with: "This meeting focused on..."', 200)
        except Exception:
            return self._extractive(full_text, 5)

    # ── Insights ──────────────────────────────────────────────────────────
    def insights(self, full_text: str, spk_data: Dict, model: str) -> Dict:
        try:
            raw = self._ollama(
                full_text, model,
                "Rate this meeting on three dimensions. Respond in EXACT format:\n"
                "CLARITY: <Good/Average/Poor> — <one reason>\n"
                "DECISIONS: <High/Medium/Low> — <one reason>\n"
                "COLLABORATION: <High/Medium/Low> — <one reason>",
                80)
            result = {"clarity": "Average", "decisions": "Medium",
                      "collaboration": "Medium",
                      "clarity_reason": "", "decisions_reason": "",
                      "collaboration_reason": ""}
            for line in raw.splitlines():
                up = line.upper()
                if up.startswith("CLARITY:"):
                    parts = line.split(":", 1)[1].split("—")
                    result["clarity"] = parts[0].strip()
                    if len(parts) > 1: result["clarity_reason"] = parts[1].strip()
                elif up.startswith("DECISIONS:"):
                    parts = line.split(":", 1)[1].split("—")
                    result["decisions"] = parts[0].strip()
                    if len(parts) > 1: result["decisions_reason"] = parts[1].strip()
                elif up.startswith("COLLABORATION:"):
                    parts = line.split(":", 1)[1].split("—")
                    result["collaboration"] = parts[0].strip()
                    if len(parts) > 1: result["collaboration_reason"] = parts[1].strip()
            return result
        except Exception:
            n = len(spk_data)
            return {"clarity": "Average", "decisions": "Medium",
                    "collaboration": "High" if n > 1 else "Low",
                    "clarity_reason": "", "decisions_reason": "",
                    "collaboration_reason": ""}


# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_excel(session: Dict, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook(); ws = wb.active; ws.title = "Meeting Analysis"

    def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr(r, c, text, fg="FFFFFF", bg="1E2D6B", sz=11, bold=True, span=1):
        if span > 1:
            ws.merge_cells(start_row=r, start_column=c,
                           end_row=r, end_column=c + span - 1)
        cell = ws.cell(r, c, text)
        cell.font      = Font(name="Calibri", bold=bold, size=sz, color=fg)
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border    = bdr()
        return cell

    spk_data        = session.get("speaker_data", {})
    n_spk           = len(spk_data)
    timeline_data   = session.get("timeline_data", [])
    timeline_topics = session.get("timeline_topics", [])
    spk_summaries   = session.get("speaker_summaries", {})
    key_pts         = session.get("key_points", [])
    action_items    = session.get("action_items", [])
    overall         = session.get("overall_summary", "")
    insights_data   = session.get("insights", {})

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 80

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"📋  MEETING SPEECH ANALYSIS  ·  {datetime.datetime.now():%Y-%m-%d %H:%M}"
    c.font      = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    c.fill      = fill("1E2D6B")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    row = 2

    def section(title, color="1E2D6B"):
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        hdr(row, 1, f"  {title}", bg=color, sz=12, span=2)
        ws.row_dimensions[row].height = 28
        row += 1

    def data_row(label, value, label_bg="F0F4FA", val_bg="FFFFFF"):
        nonlocal row
        c1 = ws.cell(row, 1, label)
        c1.font      = Font(name="Calibri", bold=True, size=10, color="1A2035")
        c1.fill      = fill(label_bg); c1.border = bdr()
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2 = ws.cell(row, 2, value)
        c2.font      = Font(name="Calibri", size=10, color="1A2035")
        c2.fill      = fill(val_bg); c2.border = bdr()
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = max(20, min(80, len(str(value)) // 3 + 16))
        row += 1

    section("📅  MEETING TIMELINE", "1E2D6B")
    for b, topic in zip(timeline_data, timeline_topics):
        data_row(f"[{fmt_time(b['start'])} – {fmt_time(b['end'])}]", topic)
    if not timeline_data:
        data_row("Timeline", "Not available")

    section("👥  SPEAKERS IDENTIFIED", "1971C2")
    for i in range(n_spk):
        data_row(f"Speaker {i + 1}", "Unknown / Not named")

    section("🗣  SPEAKER-WISE SUMMARY", "2F9E44")
    SPK_CLR = ["E7F5FF","EBFBEE","FFF3BF","F3F0FF","E6FCF5","FFE3EC"]
    for i, summ in spk_summaries.items():
        bg = SPK_CLR[i % len(SPK_CLR)]
        data_row(f"Speaker {i+1} — Discussed",    summ.get("discussed",""),    bg, bg)
        data_row(f"Speaker {i+1} — Contribution", summ.get("contribution",""), bg, bg)
        data_row(f"Speaker {i+1} — Tone",         summ.get("tone",""),         bg, bg)

    section("📌  KEY POINTS EXTRACTED", "E67700")
    for j, pt in enumerate(key_pts, 1):
        data_row(f"Point {j}", pt)

    section("📊  ACTION ITEMS", "7048E8")
    for j, item in enumerate(action_items, 1):
        data_row(f"Task {j}", item)

    section("🧠  OVERALL MEETING SUMMARY", "0C8599")
    data_row("Summary", overall, "E6FCF5", "F0FFF4")

    section("📈  INSIGHTS & OBSERVATIONS", "C2255C")
    data_row("Communication Clarity",
             f"{insights_data.get('clarity','—')}  —  {insights_data.get('clarity_reason','')}",
             "FFE3EC", "FFF5F5")
    data_row("Decision Effectiveness",
             f"{insights_data.get('decisions','—')}  —  {insights_data.get('decisions_reason','')}",
             "FFE3EC", "FFF5F5")
    data_row("Collaboration Level",
             f"{insights_data.get('collaboration','—')}  —  {insights_data.get('collaboration_reason','')}",
             "FFE3EC", "FFF5F5")

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"[Excel] Saved → {path}")


# ─── GUI Helpers ──────────────────────────────────────────────────────────────
def _lbl(parent, text, font_size=11, bold=False, color=TEXT,
         wrap=0, anchor="w", justify="left") -> ctk.CTkLabel:
    weight = "bold" if bold else "normal"
    return ctk.CTkLabel(parent, text=text,
                        font=ctk.CTkFont("Segoe UI", font_size, weight),
                        text_color=color, wraplength=wrap,
                        anchor=anchor, justify=justify)

def _badge(parent, text: str, fg: str, bg: str) -> ctk.CTkLabel:
    return ctk.CTkLabel(parent, text=f"  {text}  ",
                        font=ctk.CTkFont("Segoe UI", 10, "bold"),
                        text_color=fg, fg_color=bg, corner_radius=12)


# ─── Main App ─────────────────────────────────────────────────────────────────
class SpeakSenseApp(ctk.CTk):

    def __init__(self):
        super().__init__()
        self.title("SpeakSense  ·  Meeting Speech Analyser")
        self.geometry("1280x900")
        self.minsize(1000, 680)
        self.configure(fg_color=BG)

        self._recording    = False
        self._frames: List = []
        self._timer_sec    = 0
        self._timer_job    = None
        self._wavedata     = np.zeros(320)
        self._processor: Optional[AudioProcessor] = None
        self._summarizer   = Summarizer()
        self._session: Dict = {}

        self._build_topbar()
        self._build_controls()
        self._build_waveform()
        self._build_results_area()
        self._waveform_loop()

    # ══════════════════════════════════════════════════════════════════════
    #  TOP BAR
    # ══════════════════════════════════════════════════════════════════════
    def _build_topbar(self):
        bar = ctk.CTkFrame(self, fg_color=TOPBAR_BG, height=56, corner_radius=0)
        bar.pack(fill="x"); bar.pack_propagate(False)

        ctk.CTkLabel(bar, text="🎙  SpeakSense",
                     font=ctk.CTkFont("Segoe UI", 21, "bold"),
                     text_color="#FFFFFF").pack(side="left", padx=24)
        ctk.CTkLabel(bar, text="Meeting Speech Analyser",
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color="#A5B4FC").pack(side="left")

        right = ctk.CTkFrame(bar, fg_color="transparent")
        right.pack(side="right", padx=20)
        self._timer_lbl = ctk.CTkLabel(right, text="00:00",
                                       font=ctk.CTkFont("Consolas", 22, "bold"),
                                       text_color="#FFFFFF")
        self._timer_lbl.pack(side="right", padx=(12, 0))
        self._status_lbl = ctk.CTkLabel(right, text="● READY",
                                        font=ctk.CTkFont("Consolas", 11, "bold"),
                                        text_color="#A5B4FC")
        self._status_lbl.pack(side="right")

    # ══════════════════════════════════════════════════════════════════════
    #  CONTROLS  (Language option removed)
    # ══════════════════════════════════════════════════════════════════════
    def _build_controls(self):
        bar = ctk.CTkFrame(self, fg_color=SURFACE, height=62,
                           corner_radius=0, border_width=1, border_color=BORDER)
        bar.pack(fill="x"); bar.pack_propagate(False)

        def lbl(t):
            ctk.CTkLabel(bar, text=t, font=ctk.CTkFont("Consolas", 9),
                         text_color=MUTED).pack(side="left", padx=(14, 3))

        def omenu(var, values, w=115):
            ctk.CTkOptionMenu(bar, variable=var, values=values,
                              width=w, height=30,
                              fg_color=RAISED, button_color=BORDER,
                              text_color=TEXT, button_hover_color=ACCENT,
                              dropdown_fg_color=SURFACE, dropdown_text_color=TEXT,
                              font=ctk.CTkFont("Consolas", 11)
                              ).pack(side="left", padx=(0, 10))

        # ── Whisper model ──────────────────────────────────────────────────
        lbl("WHISPER"); self._w_var = ctk.StringVar(value="base")
        omenu(self._w_var, ["tiny", "base", "small", "medium", "large"])

        # ── Ollama model ───────────────────────────────────────────────────
        lbl("OLLAMA"); self._o_var = ctk.StringVar(value="llama3")
        omenu(self._o_var, ["llama3","llama3.1","mistral","gemma2","phi3","qwen2"], 125)

        # ── Number of speakers ─────────────────────────────────────────────
        lbl("SPEAKERS"); self._spk_var = ctk.StringVar(value="2")
        omenu(self._spk_var, ["1","2","3","4","5","6"], 70)

        # ── Separator ─────────────────────────────────────────────────────
        ctk.CTkFrame(bar, fg_color=BORDER, width=1, height=36,
                     corner_radius=0).pack(side="left", padx=10)

        def btn(text, fg, hv, cmd, state="normal", txt_col="#FFFFFF"):
            b = ctk.CTkButton(bar, text=text, width=155, height=34,
                              fg_color=fg, hover_color=hv,
                              text_color=txt_col, state=state,
                              font=ctk.CTkFont("Segoe UI", 12, "bold"),
                              command=cmd)
            b.pack(side="left", padx=3)
            return b

        self._btn_start = btn("⏺  Start Recording", RED,    "#B91C1C", self._start)
        self._btn_stop  = btn("⏹  Stop & Analyse",  BORDER, "#94A3B8", self._stop,
                              state="disabled", txt_col=TEXT)
        self._btn_load  = btn("📂  Load File",       ACCENT, "#3730A3", self._load_file)
        self._btn_xl    = btn("📊  Save Excel",       GREEN,  "#15803D", self._save_excel,
                              state="disabled")

        # Progress
        self._prog_frame = ctk.CTkFrame(bar, fg_color="transparent")
        self._prog_lbl   = ctk.CTkLabel(self._prog_frame, text="",
                                        font=ctk.CTkFont("Consolas", 10),
                                        text_color=ACCENT)
        self._prog_lbl.pack(side="left", padx=(12, 6))
        self._prog_bar = ctk.CTkProgressBar(self._prog_frame, width=170, height=8,
                                            fg_color=BORDER, progress_color=ACCENT)
        self._prog_bar.set(0); self._prog_bar.pack(side="left")

    # ══════════════════════════════════════════════════════════════════════
    #  WAVEFORM
    # ══════════════════════════════════════════════════════════════════════
    def _build_waveform(self):
        f = ctk.CTkFrame(self, fg_color=SURFACE, height=52,
                         corner_radius=0, border_width=1, border_color=BORDER)
        f.pack(fill="x"); f.pack_propagate(False)
        self._wave_cv = tk.Canvas(f, bg=SURFACE, highlightthickness=0, height=52)
        self._wave_cv.pack(fill="both", expand=True, padx=12, pady=4)

    def _waveform_loop(self):
        try:
            c = self._wave_cv; W = c.winfo_width() or 960; H = 52
            c.delete("all")
            if self._recording:
                n = len(self._wavedata); bw = W / n
                for i, v in enumerate(self._wavedata):
                    bh = max(2, v * H * 4); x = i * bw
                    r  = int(60 + v * 195); gb = int(max(0, 90 - v * 90))
                    c.create_rectangle(x, H/2 - bh/2, x + bw - 1, H/2 + bh/2,
                                       fill=f"#{r:02x}{gb:02x}{gb:02x}", outline="")
            else:
                c.create_text(W/2, H/2,
                              text="── live waveform appears during recording ──",
                              fill=MUTED, font=("Consolas", 10))
        except Exception:
            pass
        self.after(40, self._waveform_loop)

    # ══════════════════════════════════════════════════════════════════════
    #  RESULTS AREA
    # ══════════════════════════════════════════════════════════════════════
    def _build_results_area(self):
        outer = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        outer.pack(fill="both", expand=True)

        self._scroll = ctk.CTkScrollableFrame(
            outer, fg_color=BG, corner_radius=0,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color=ACCENT)
        self._scroll.pack(fill="both", expand=True)
        self._scroll.grid_columnconfigure(0, weight=1)

        self._placeholder = ctk.CTkLabel(
            self._scroll,
            text=(
                "🎙  Press  ⏺ Start Recording  or  📂 Load File  to begin.\n\n"
                "Your full meeting analysis will appear here:\n\n"
                "  📅 Timeline  ·  👥 Speakers  ·  🗣 Speaker-wise Summary\n"
                "  📌 Key Points  ·  📊 Action Items  ·  🧠 Overall Summary  ·  📈 Insights"
            ),
            font=ctk.CTkFont("Segoe UI", 14),
            text_color=MUTED, justify="center")
        self._placeholder.pack(expand=True, pady=120)

    def _clear_results(self):
        for w in self._scroll.winfo_children():
            w.destroy()

    # ── Card / row helpers ────────────────────────────────────────────────
    def _make_card(self, emoji: str, title: str,
                   accent_fg: str, accent_bg: str) -> ctk.CTkFrame:
        outer = ctk.CTkFrame(self._scroll, fg_color=CARD, corner_radius=12,
                             border_width=1, border_color=BORDER)
        outer.pack(fill="x", padx=16, pady=(0, 12))

        hdr = ctk.CTkFrame(outer, fg_color=accent_bg, corner_radius=8, height=38)
        hdr.pack(fill="x", padx=8, pady=(8, 0))
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text=f"{emoji}  {title}",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=accent_fg).pack(side="left", padx=14, pady=6)

        body = ctk.CTkFrame(outer, fg_color="transparent")
        body.pack(fill="x", padx=8, pady=(6, 8))
        return body

    def _row_pair(self, parent, label: str, value: str,
                  label_fg=MUTED, val_fg=TEXT, val_bg=RAISED, val_wrap=900):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=2)
        row.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(row, text=label,
                     font=ctk.CTkFont("Segoe UI", 10, "bold"),
                     text_color=label_fg, width=200, anchor="w"
                     ).grid(row=0, column=0, padx=(4, 8), sticky="nw", pady=4)

        vf = ctk.CTkFrame(row, fg_color=val_bg, corner_radius=6)
        vf.grid(row=0, column=1, sticky="ew", padx=(0, 4), pady=2)
        ctk.CTkLabel(vf, text=value,
                     font=ctk.CTkFont("Segoe UI", 11),
                     text_color=val_fg, wraplength=val_wrap,
                     anchor="nw", justify="left"
                     ).pack(anchor="nw", padx=10, pady=6, fill="x")

    # ══════════════════════════════════════════════════════════════════════
    #  RENDER ALL SECTIONS
    # ══════════════════════════════════════════════════════════════════════
    def _render_results(self):
        self._clear_results()
        sess = self._session

        ctk.CTkLabel(self._scroll, text="", height=8).pack()

        # ── 1. TIMELINE ───────────────────────────────────────────────────
        body = self._make_card("📅", "MEETING TIMELINE", "#1E2D6B", LIGHT_BLU)
        tl   = sess.get("timeline_data", [])
        tops = sess.get("timeline_topics", [])
        if tl:
            for b, topic in zip(tl, tops):
                self._row_pair(body,
                               f"[{fmt_time(b['start'])} – {fmt_time(b['end'])}]",
                               topic, label_fg=ACCENT)
        else:
            self._row_pair(body, "Status", "Timeline will appear after processing.")

        # ── 2. SPEAKERS IDENTIFIED ────────────────────────────────────────
        body     = self._make_card("👥", "SPEAKERS IDENTIFIED", "#1971C2", "#DBEAFE")
        n        = sess.get("n_speakers", 0)
        spk_data = sess.get("speaker_data", {})
        if n:
            for i in range(n):
                d    = spk_data.get(i, {})
                dur  = d.get("duration", 0)
                cnt  = d.get("count", 0)
                info = (f"Unknown  ·  spoke {fmt_time(dur)} total  "
                        f"·  {cnt} segment{'s' if cnt != 1 else ''}")
                row = ctk.CTkFrame(body, fg_color="transparent")
                row.pack(fill="x", pady=3)
                _badge(row, f"SPK {i+1}", spk_fg(i), spk_bg(i)).pack(side="left", padx=(4, 10))
                ctk.CTkLabel(row, text=info,
                             font=ctk.CTkFont("Segoe UI", 11),
                             text_color=TEXT).pack(side="left")
        else:
            self._row_pair(body, "Status", "Speakers will appear after processing.")

        # ── 3. SPEAKER-WISE SUMMARY ───────────────────────────────────────
        body     = self._make_card("🗣", "SPEAKER-WISE SUMMARY", "#2F9E44", "#DCFCE7")
        spk_sums = sess.get("speaker_summaries", {})
        if spk_sums:
            for i, summ in spk_sums.items():
                sh = ctk.CTkFrame(body, fg_color=spk_bg(i), corner_radius=8,
                                  border_width=1, border_color=spk_bd(i), height=30)
                sh.pack(fill="x", pady=(6, 2)); sh.pack_propagate(False)
                ctk.CTkLabel(sh, text=f"🔹  {spk_label(i)}",
                             font=ctk.CTkFont("Segoe UI", 11, "bold"),
                             text_color=spk_fg(i)).pack(side="left", padx=12)

                self._row_pair(body, "  Discussed",
                               summ.get("discussed", "—"), label_fg=spk_fg(i), val_bg=spk_bg(i))
                self._row_pair(body, "  Key Contribution",
                               summ.get("contribution", "—"), label_fg=spk_fg(i), val_bg=spk_bg(i))

                tone = summ.get("tone", "—")
                tone_colors = {
                    "Formal":        ("#1E40AF", "#DBEAFE"),
                    "Informal":      ("#92400E", "#FEF3C7"),
                    "Analytical":    ("#5B21B6", "#EDE9FE"),
                    "Collaborative": ("#065F46", "#D1FAE5"),
                    "Critical":      ("#9F1239", "#FFE4E6"),
                }
                tc = tone_colors.get(tone, (MUTED, RAISED))
                tone_row = ctk.CTkFrame(body, fg_color="transparent")
                tone_row.pack(fill="x", pady=2)
                ctk.CTkLabel(tone_row, text="  Tone",
                             font=ctk.CTkFont("Segoe UI", 10, "bold"),
                             text_color=spk_fg(i), width=200, anchor="w"
                             ).pack(side="left", padx=(4, 8))
                _badge(tone_row, tone, tc[0], tc[1]).pack(side="left", padx=4)
                ctk.CTkFrame(body, fg_color=BORDER, height=1).pack(
                    fill="x", pady=(8, 0), padx=4)
        else:
            self._row_pair(body, "Status", "Speaker summaries generating…")

        # ── 4. KEY POINTS ─────────────────────────────────────────────────
        body    = self._make_card("📌", "KEY POINTS EXTRACTED", "#E67700", "#FEF9C3")
        kp_list = sess.get("key_points", [])
        if kp_list:
            for j, pt in enumerate(kp_list, 1):
                row = ctk.CTkFrame(body, fg_color=RAISED, corner_radius=6)
                row.pack(fill="x", pady=3, padx=4)
                ctk.CTkLabel(row, text=f"  Point {j}",
                             font=ctk.CTkFont("Consolas", 10, "bold"),
                             text_color=AMBER, width=80, anchor="w"
                             ).pack(side="left", padx=(8, 6), pady=6)
                ctk.CTkLabel(row, text=pt,
                             font=ctk.CTkFont("Segoe UI", 11),
                             text_color=TEXT, wraplength=850,
                             anchor="nw", justify="left"
                             ).pack(side="left", padx=4, pady=6, fill="x")
        else:
            self._row_pair(body, "Status", "Key points generating…")

        # ── 5. ACTION ITEMS ───────────────────────────────────────────────
        body    = self._make_card("📊", "ACTION ITEMS", "#7048E8", "#EDE9FE")
        ai_list = sess.get("action_items", [])
        if ai_list:
            for j, item in enumerate(ai_list, 1):
                row = ctk.CTkFrame(body, fg_color=LIGHT_PUR, corner_radius=6,
                                   border_width=1, border_color="#DDD6FE")
                row.pack(fill="x", pady=3, padx=4)
                ctk.CTkLabel(row, text=f"  ✅  Task {j}",
                             font=ctk.CTkFont("Segoe UI", 10, "bold"),
                             text_color=PURPLE, width=90, anchor="w"
                             ).pack(side="left", padx=(8, 6), pady=6)
                ctk.CTkLabel(row, text=item,
                             font=ctk.CTkFont("Segoe UI", 11),
                             text_color=TEXT, wraplength=840,
                             anchor="nw", justify="left"
                             ).pack(side="left", padx=4, pady=6, fill="x")
        else:
            self._row_pair(body, "Status", "Action items generating…")

        # ── 6. OVERALL SUMMARY ────────────────────────────────────────────
        body    = self._make_card("🧠", "OVERALL MEETING SUMMARY", "#0C8599", LIGHT_TEL)
        overall = sess.get("overall_summary", "")
        if overall:
            frame = ctk.CTkFrame(body, fg_color="#F0FDFA", corner_radius=8,
                                 border_width=1, border_color="#99F6E4")
            frame.pack(fill="x", padx=4, pady=4)
            ctk.CTkLabel(frame, text=overall,
                         font=ctk.CTkFont("Segoe UI", 12),
                         text_color="#0F4C5C", wraplength=1060,
                         justify="left", anchor="nw"
                         ).pack(anchor="nw", padx=14, pady=12, fill="x")
        else:
            self._row_pair(body, "Status", "Overall summary generating…")

        # ── 7. INSIGHTS ───────────────────────────────────────────────────
        body = self._make_card("📈", "INSIGHTS & OBSERVATIONS", "#C2255C", "#FFE4E6")
        ins  = sess.get("insights", {})
        if ins:
            self._insight_row(body, "Communication Clarity",
                              ins.get("clarity","—"), ins.get("clarity_reason",""))
            self._insight_row(body, "Decision Effectiveness",
                              ins.get("decisions","—"), ins.get("decisions_reason",""))
            self._insight_row(body, "Collaboration Level",
                              ins.get("collaboration","—"), ins.get("collaboration_reason",""))
        else:
            self._row_pair(body, "Status", "Insights generating…")

        ctk.CTkLabel(self._scroll, text="", height=20).pack()

    def _insight_row(self, parent, label: str, rating: str, reason: str):
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=4)
        ctk.CTkLabel(row, text=label,
                     font=ctk.CTkFont("Segoe UI", 11, "bold"),
                     text_color=TEXT, width=220, anchor="w"
                     ).pack(side="left", padx=(4, 8))

        color_map = {
            "Good":    ("#166534","#DCFCE7"),
            "High":    ("#166534","#DCFCE7"),
            "Average": ("#92400E","#FEF3C7"),
            "Medium":  ("#92400E","#FEF3C7"),
            "Poor":    ("#991B1B","#FEE2E2"),
            "Low":     ("#991B1B","#FEE2E2"),
        }
        fg_c, bg_c = color_map.get(rating, (MUTED, RAISED))
        _badge(row, rating, fg_c, bg_c).pack(side="left", padx=4)
        if reason:
            ctk.CTkLabel(row, text=f"—  {reason}",
                         font=ctk.CTkFont("Segoe UI", 10),
                         text_color=MUTED).pack(side="left", padx=8)

    # ══════════════════════════════════════════════════════════════════════
    #  PIPELINE
    # ══════════════════════════════════════════════════════════════════════
    def _pipeline(self, wav_path: str):
        try:
            wm    = self._w_var.get()
            om    = self._o_var.get()
            n_spk = int(self._spk_var.get())

            self._emit("progress", (0.08, "Loading Whisper…"))
            if self._processor is None or self._processor.model_size != wm:
                self._processor = AudioProcessor(wm)

            self._emit("progress", (0.18, "Transcribing audio…"))
            data = self._processor.process(wav_path, n_speakers=n_spk)
            self._session.update(data)
            self._session["n_speakers"] = len(data["speaker_data"])

            self._emit("progress", (0.30, "Building timeline…"))
            buckets = self._summarizer.build_timeline(data["segments"], data["duration"])
            topics  = self._summarizer.timeline_topics(buckets, om)
            self._session["timeline_data"]   = buckets
            self._session["timeline_topics"] = topics

            self._emit("progress", (0.42, "Analysing speakers…"))
            spk_summaries: Dict[int, Dict] = {}
            for i, info in data["speaker_data"].items():
                spk_summaries[i] = self._summarizer.speaker_summary(info["text"], i, om)
                p = 0.42 + 0.18 * ((i + 1) / max(1, len(data["speaker_data"])))
                self._emit("progress", (p, f"Speaker {i+1} analysed…"))
            self._session["speaker_summaries"] = spk_summaries

            self._emit("progress", (0.62, "Extracting key points…"))
            self._session["key_points"] = self._summarizer.key_points(data["full_text"], om)

            self._emit("progress", (0.72, "Identifying action items…"))
            self._session["action_items"] = self._summarizer.action_items(data["full_text"], om)

            self._emit("progress", (0.82, "Writing overall summary…"))
            self._session["overall_summary"] = self._summarizer.overall_summary(
                data["full_text"], spk_summaries, om)

            self._emit("progress", (0.91, "Generating insights…"))
            self._session["insights"] = self._summarizer.insights(
                data["full_text"], data["speaker_data"], om)

            self._emit("progress", (0.95, "Saving Excel…"))
            ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            Path("exports").mkdir(exist_ok=True)
            xls = f"exports/SpeakSense_{ts}.xlsx"
            export_excel(self._session, xls)

            self._emit("complete", xls)

        except Exception as e:
            import traceback; traceback.print_exc()
            self._emit("error", str(e))

    # ══════════════════════════════════════════════════════════════════════
    #  EVENT BUS
    # ══════════════════════════════════════════════════════════════════════
    def _emit(self, ev: str, data):
        self.after(0, lambda: self._handle(ev, data))

    def _handle(self, ev: str, data):
        if ev == "progress":
            pct, lbl = data
            self._prog_bar.set(pct); self._prog_lbl.configure(text=lbl)
        elif ev == "complete":
            xls = data
            self._render_results()
            self._show_progress(False)
            self._btn_start.configure(state="normal", fg_color=RED)
            self._btn_xl.configure(state="normal")
            self._set_status("✅  DONE", GREEN)
            mb.showinfo("Analysis Complete 🎉",
                        f"Full meeting analysis ready!\n\nExcel saved to:\n{xls}")
        elif ev == "error":
            self._show_progress(False)
            self._btn_start.configure(state="normal", fg_color=RED)
            self._set_status("✗  ERROR", RED)
            mb.showerror("Error", str(data))

    # ══════════════════════════════════════════════════════════════════════
    #  RECORDING
    # ══════════════════════════════════════════════════════════════════════
    def _start(self):
        try:
            import sounddevice as sd
            self._recording = True; self._frames = []; self._timer_sec = 0
            self._set_status("🔴  RECORDING", RED)
            self._btn_start.configure(state="disabled")
            self._btn_stop.configure(state="normal", fg_color=RED, text_color="#FFFFFF")
            self._show_progress(False); self._tick()

            def cb(indata, frames, t, status):
                if self._recording:
                    self._frames.append(indata.copy())
                    v = np.linalg.norm(indata) / len(indata)
                    self._wavedata = np.roll(self._wavedata, -1)
                    self._wavedata[-1] = v

            self._stream = sd.InputStream(samplerate=16000, channels=1,
                                          dtype="float32", callback=cb)
            self._stream.start()
        except Exception as e:
            mb.showerror("Recording Error", str(e))

    def _stop(self):
        try:
            self._recording = False
            if hasattr(self, "_stream"):
                self._stream.stop(); self._stream.close()
            if self._timer_job: self.after_cancel(self._timer_job)
            self._btn_start.configure(state="normal", fg_color=RED)
            self._btn_stop.configure(state="disabled", fg_color=BORDER, text_color=TEXT)
            self._set_status("⚙  PROCESSING", AMBER)
            self._show_progress(True, 0.05, "Preparing audio…")

            import soundfile as sf
            Path("temp").mkdir(exist_ok=True)
            wav = f"temp/{uuid.uuid4().hex[:8]}.wav"
            sf.write(wav, np.concatenate(self._frames, axis=0), 16000)
            threading.Thread(target=self._pipeline, args=(wav,), daemon=True).start()
        except Exception as e:
            mb.showerror("Error", str(e))

    def _load_file(self):
        path = fd.askopenfilename(
            title="Select Audio File",
            filetypes=[("Audio", "*.wav *.mp3 *.m4a *.flac *.ogg"), ("All", "*.*")])
        if not path: return
        self._set_status("⚙  LOADING", AMBER)
        self._show_progress(True, 0.05, "Converting audio…")
        try:
            wav = self._to_wav(path)
        except Exception as e:
            mb.showerror("Error", f"Could not load:\n{e}")
            self._show_progress(False); return
        threading.Thread(target=self._pipeline, args=(wav,), daemon=True).start()

    def _to_wav(self, path: str) -> str:
        if path.lower().endswith(".wav"): return path
        try:
            from pydub import AudioSegment
        except ImportError:
            raise RuntimeError("pip install pydub")
        Path("temp").mkdir(exist_ok=True)
        out = f"temp/{uuid.uuid4().hex[:8]}.wav"
        (AudioSegment.from_file(path)
         .set_channels(1).set_frame_rate(16000).export(out, format="wav"))
        return out

    def _save_excel(self):
        if not self._session.get("segments"):
            mb.showinfo("Nothing to Save", "Process a recording first."); return
        path = fd.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"SpeakSense_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx")
        if not path: return
        try:
            export_excel(self._session, path)
            mb.showinfo("Saved!", f"Saved to:\n{path}")
        except Exception as e:
            mb.showerror("Failed", str(e))

    # ── Helpers ───────────────────────────────────────────────────────────
    def _set_status(self, text: str, color: str):
        self._status_lbl.configure(text=text, text_color=color)

    def _show_progress(self, show: bool, pct: float = 0.0, lbl: str = ""):
        if show:
            self._prog_bar.set(pct); self._prog_lbl.configure(text=lbl)
            self._prog_frame.pack(side="left", padx=(10, 0))
        else:
            self._prog_frame.pack_forget()

    def _tick(self):
        if self._recording:
            self._timer_sec += 1
            m = self._timer_sec // 60; s = self._timer_sec % 60
            self._timer_lbl.configure(text=f"{m:02d}:{s:02d}")
            self._timer_job = self.after(1000, self._tick)


# ─── Entry Point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SpeakSenseApp()
    app.mainloop()