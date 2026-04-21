import os, re, sys, wave, uuid, time, queue, threading, datetime
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
import customtkinter as ctk

# ── Light / White theme ───────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ─── Palette (WHITE / LIGHT) ──────────────────────────────────────────────────
BG = "#F4F7FB" # page background — very light blue-grey
SURFACE = "#FFFFFF" # panels / topbar
CARD = "#FFFFFF" # main card background
RAISED = "#EDF1F7" # slightly raised surfaces, option-menus
BORDER = "#CBD5E1" # dividers, borders
TEXT = "#1E293B" # primary text (near-black)
MUTED = "#64748B" # secondary / placeholder text
RED = "#E53E3E" # record button
GREEN = "#16A34A" # done / excel button
BLUE = "#2563EB" # logo, progress
PURPLE = "#7C3AED"
AMBER = "#D97706"
TEAL = "#0D9488"
DARK_GREEN = "#14532D" # excel button text

# ── Speaker accent colours stay vivid; backgrounds & borders go LIGHT ─────────
SPEAKER_FG = ["#0369A1", "#15803D", "#B45309", "#9D174D", "#6D28D9", "#0F766E"]
SPEAKER_BG = ["#EFF6FF", "#F0FDF4", "#FFFBEB", "#FFF0F6", "#F5F3FF", "#F0FDFA"]
SPEAKER_BD = ["#BFDBFE", "#BBF7D0", "#FDE68A", "#FBCFE8", "#DDD6FE", "#99F6E4"]

def spk_fg(i): return SPEAKER_FG[i % len(SPEAKER_FG)]
def spk_bg(i): return SPEAKER_BG[i % len(SPEAKER_BG)]
def spk_bd(i): return SPEAKER_BD[i % len(SPEAKER_BD)]
def spk_label(i): return f"Speaker {i+1}"
def fmt_time(s):
 m = int(s)//60; sec = int(s)%60
 return f"{m:02d}:{sec:02d}"

# ─── Audio Processor ──────────────────────────────────────────────────────────
class AudioProcessor:
    def real_diarization(self, wav_path):
        from pyannote.audio import Pipeline
        pipeline = Pipeline.from_pretrained("pyannote/speaker-diarization")
        diarization = pipeline(wav_path)

        speakers = []
        for turn, _, speaker in diarization.itertracks(yield_label=True):
            speakers.append({
                "start": turn.start,
                "end": turn.end,
                "speaker": speaker
            })

        return speakers

    def __init__(self, model_size="medium"):
        import whisper
        self.model = whisper.load_model(model_size)
        self.model_size = model_size

    def process(self, wav_path: str) -> Dict:
        r = self.model.transcribe(
            str(wav_path),
            word_timestamps=True,
            verbose=False,
            fp16=False,
            language="hi",
            task="transcribe"
        )

        segs = [{
            "id": s["id"],
            "start": round(s["start"], 2),
            "end": round(s["end"], 2),
            "text": s["text"].strip(),
            "speaker": 0
        } for s in r["segments"]]

        diar = self.real_diarization(wav_path)

        for seg in segs:
            for d in diar:
                if d["start"] <= seg["start"] <= d["end"]:
                    seg["speaker"] = hash(d["speaker"]) % 10
                    break

        spk_data = self._build_speakers(segs)

        return {
            "segments": segs,
            "speaker_data": spk_data,
            "full_text": " ".join(s["text"] for s in segs),
            "duration": segs[-1]["end"] if segs else 0,
        }

    def _build_speakers(self, segs) -> Dict[int, Dict]:
        data = {}
        for s in segs:
            idx = s["speaker"]
            if idx not in data:
                data[idx] = {
                    "text": "",
                    "segments": [],
                    "duration": 0.0,
                    "count": 0,
                    "first": s["start"],
                    "last": s["end"]
                }

            d = data[idx]
            d["text"] += (" " + s["text"]).lstrip()
            d["segments"].append({
                "start": s["start"],
                "end": s["end"],
                "text": s["text"]
            })
            d["duration"] += round(s["end"] - s["start"], 2)
            d["count"] += 1
            d["last"] = max(d["last"], s["end"])

        for d in data.values():
            d["text"] = d["text"].strip()

        return dict(sorted(data.items()))

# ─── Summarizer ───────────────────────────────────────────────────────────────
import re
from collections import Counter
from typing import Dict


def spk_label(idx):
    return f"Speaker {idx + 1}"


class Summarizer:
    STOP = {
        "the","a","an","in","on","at","to","for","of","and","or","but","is","are",
        "was","were","be","been","it","this","that","with","i","we","you","he",
        "she","they","so","as","by","from","not","have","has","had","do","did"
    }

    def ollama_sum(self, text, model="llama3", max_w=160, hint=""):
        import ollama

        prompt = (
            f"Summarise in {max_w} words or fewer. Concise, accurate. {hint}\n\n"
            f"TEXT:\n{text[:4000]}\n\nSUMMARY:"
        )

        response = ollama.chat(
            model=model,
            messages=[{"role": "user", "content": prompt}]
        )

        return response["message"]["content"].strip()

    def bart_sum(self, text):
        try:
            return self.ollama_sum(text, "llama3", 150)
        except Exception:
            return self._extractive(text, 4)

    def _extractive(self, text, n=4):
        # Split sentences
        sents = [
            s.strip()
            for s in re.split(r"(?<=[.!?])\s+", text)
            if len(s.split()) >= 5
        ]

        if not sents:
            return text[:400]

        # Word frequency
        freq = Counter(re.findall(r"\b\w+\b", text.lower()))

        scored = []

        for s in sents:
            words = [
                w for w in re.findall(r"\b\w+\b", s.lower())
                if w not in self.STOP
            ]

            score = sum(freq.get(w, 0) for w in words) / (len(words) + 1)
            scored.append((score, s))

        # Sort by score
        scored.sort(reverse=True)

        # Preserve original order
        top = sorted(scored[:n], key=lambda x: sents.index(x[1]))

        return " ".join(x[1] for x in top)

    def speaker_sum(self, text, idx, model="llama3"):
        if len(text.split()) < 15:
            return text.strip()

        try:
            return self.ollama_sum(
                text,
                model,
                80,
                f"Focus only on what {spk_label(idx)} said."
            )
        except Exception:
            return self._extractive(text, 2)

    def best_overall(self, full_text, spk_summaries: Dict[int, str], model="llama3"):
        combined = "\n".join(
            f"[{spk_label(i)}]: {s}"
            for i, s in spk_summaries.items()
        )

        try:
            return self.ollama_sum(
                combined,
                model,
                200,
                "Extract the most important points from ALL speakers. "
                "Write a unified meeting/conversation summary covering key decisions, "
                "topics, and agreements. Do NOT attribute to speakers—write as a whole."
            )
        except Exception:
            return self._extractive(full_text, 6)

# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_excel(session: Dict, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Session Results"

    def fill(h):
        return PatternFill("solid", fgColor=h.lstrip("#"))

    def bdr():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    spk_data = session["speaker_data"]
    spk_summaries = session["speaker_summaries"]
    segments = session["segments"]
    overall = session["overall_summary"]
    n_spk = len(spk_data)

    ws.column_dimensions["A"].width = 18
    for i in range(n_spk):
        ws.column_dimensions[get_column_letter(2+i)].width = 45

    last_col = get_column_letter(1 + n_spk)

    # Header
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"SpeakSense Recording — {datetime.datetime.now():%Y-%m-%d %H:%M}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = fill("0D1B2A")
    c.alignment = Alignment(horizontal="center")

    # Column headers
    ws.cell(2,1,"TIMELINE").font = Font(bold=True)
    for i in range(n_spk):
        ws.cell(2,2+i,f"SPEAKER {i+1}").font = Font(bold=True)

    row = 3

    # Segments
    for seg in segments:
        idx = seg["speaker"]

        ws.cell(row,1,f"{fmt_time(seg['start'])}→{fmt_time(seg['end'])}")
        ws.cell(row,2+idx,seg["text"])

        row += 1

    # Speaker summaries
    ws.cell(row,1,"SPEAKER SUMMARIES").font = Font(bold=True)
    row += 1

    for idx, summ in spk_summaries.items():
        ws.cell(row,1,f"Speaker {idx+1}")
        ws.cell(row,2,summ)
        row += 1

    # Overall summary
    ws.cell(row,1,"OVERALL SUMMARY").font = Font(bold=True)
    row += 1
    ws.cell(row,1,overall)

    wb.save(path)
    print(f"[Excel] Saved → {path}")


# ─── Main App ─────────────────────────────────────────────────────────────────
class SpeakSenseApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("SpeakSense")
        self.geometry("1400x860")
        self.minsize(1000, 640)
        self.configure(fg_color=BG)

        # State
        self._recording = False
        self._frames = []
        self._timer_sec = 0
        self._timer_job = None
        self._wavedata = np.zeros(300)

        self._processor = None
        self._summarizer = Summarizer()

        self._session = {}

        # Table state (IMPORTANT FIX)
        self._row_widgets = []
        self._spk_cols = []
        self._n_speakers = 0
        self._next_seg_row = 1
        self._seg_row_start = 1
        self._seg_cells = {}

        # Build UI
        self._build_topbar()
        self._build_controls()
        self._build_waveform()
        self._build_main_table()

        self._waveform_loop()

 # ── Top bar ───────────────────────────────────────────────────────────
def _build_topbar(self):
    bar = ctk.CTkFrame(self, height=52)
    bar.pack(fill="x")

    ctk.CTkLabel(
        bar,
        text="🎙 SpeakSense",
        font=ctk.CTkFont(size=19, weight="bold"),
        text_color=BLUE
    ).pack(side="left", padx=20)

    right = ctk.CTkFrame(bar, fg_color="transparent")
    right.pack(side="right", padx=20)

    self._timer_lbl = ctk.CTkLabel(
        right,
        text="00:00",
        font=ctk.CTkFont(size=20, weight="bold")
    )
    self._timer_lbl.pack(side="right")

    self._status_lbl = ctk.CTkLabel(
        right,
        text="READY",
        text_color=MUTED
    )
    self._status_lbl.pack(side="right", padx=10)
 # ── Controls row ──────────────────────────────────────────────────────
def _build_controls(self):
    bar = ctk.CTkFrame(self, height=60)
    bar.pack(fill="x")

    # Whisper model
    self._w_var = ctk.StringVar(value="base")
    ctk.CTkOptionMenu(
        bar,
        variable=self._w_var,
        values=["tiny", "base", "small", "medium", "large"]
    ).pack(side="left", padx=10)

    # Ollama model
    self._o_var = ctk.StringVar(value="llama3")
    ctk.CTkOptionMenu(
        bar,
        variable=self._o_var,
        values=["llama3", "mistral", "gemma2"]
    ).pack(side="left", padx=10)

    # Buttons
    self._btn_start = ctk.CTkButton(
        bar, text="Start Recording", command=self._start
    )
    self._btn_start.pack(side="left", padx=10)

    self._btn_stop = ctk.CTkButton(
        bar, text="Stop", command=self._stop, state="disabled"
    )
    self._btn_stop.pack(side="left", padx=10)

    self._btn_load = ctk.CTkButton(
        bar, text="Load File", command=self._load_audio_file
    )
    self._btn_load.pack(side="left", padx=10)

    self._btn_xl = ctk.CTkButton(
        bar, text="Save Excel", command=self._save_excel, state="disabled"
    )
    self._btn_xl.pack(side="left", padx=10) 
def _build_waveform(self):
    self._wave_frame = ctk.CTkFrame(self, height=60)
    self._wave_frame.pack(fill="x")

    self._wave_canvas = tk.Canvas(self._wave_frame, height=60)
    self._wave_canvas.pack(fill="both", expand=True)
def _waveform_loop(self):
    try:
        c = self._wave_canvas
        W = c.winfo_width() or 800
        H = 60

        c.delete("all")

        if self._recording:
            n = len(self._wavedata)
            bw = W / n

            for i, v in enumerate(self._wavedata):
                bh = max(2, v * H * 3)
                x = i * bw

                c.create_rectangle(
                    x, H/2-bh/2,
                    x+bw-1, H/2+bh/2,
                    fill="red"
                )
        else:
            c.create_text(W/2, H/2, text="Waveform idle")

    except:
        pass

    self.after(50, self._waveform_loop)
def _build_main_table(self):
    self._scroll = ctk.CTkScrollableFrame(self)
    self._scroll.pack(fill="both", expand=True, padx=10, pady=10)

    self._next_seg_row = 1
def _add_segment_row(self, seg):
    row = self._next_seg_row

    # REAL timeline (FIXED)
    timeline = f"{fmt_time(seg['start'])} → {fmt_time(seg['end'])}"

    ctk.CTkLabel(self._scroll, text=timeline).grid(row=row, column=0)
    ctk.CTkLabel(self._scroll, text=f"Speaker {seg['speaker']+1}").grid(row=row, column=1)

    ctk.CTkLabel(
        self._scroll,
        text=seg["text"],
        wraplength=600,
        justify="left"
    ).grid(row=row, column=2)

    self._next_seg_row += 1
def _build_summary_section(self):
    row = self._next_seg_row

    ctk.CTkLabel(
        self._scroll,
        text="Speaker Summaries",
        font=ctk.CTkFont(size=14, weight="bold")
    ).grid(row=row, column=0, columnspan=3)

    row += 1

    self._spk_sum_labels = {}

    for i in range(self._n_speakers):
        lbl = ctk.CTkLabel(self._scroll, text=f"Speaker {i+1} summary...")
        lbl.grid(row=row, column=0, columnspan=3)

        self._spk_sum_labels[i] = lbl
        row += 1

    self._overall_lbl = ctk.CTkLabel(
        self._scroll,
        text="Overall summary..."
    )
    self._overall_lbl.grid(row=row, column=0, columnspan=3)

    self._next_seg_row = row + 1

    def _emit(self, ev, data):
        self.after(0, lambda: self._handle(ev, data))

    def _handle(self, ev, data):
        if ev == "progress":
            pct, label = data
            self._prog_bar.set(pct)
            self._prog_lbl.configure(text=label)

        elif ev == "transcript_ready":
            self._rebuild_table()
            for seg in data["segments"]:
                self._add_segment_row(seg)
            self._build_summary_section()

        elif ev == "spk_summary":
            i, summ = data
            if hasattr(self, "_spk_sum_labels"):
                self._spk_sum_labels[i].configure(text=summ)

        elif ev == "complete":
            overall, xls = data

            if hasattr(self, "_overall_lbl"):
                self._overall_lbl.configure(text=overall)

            self._btn_start.configure(state="normal")
            self._btn_xl.configure(state="normal")

            mb.showinfo("Done", f"Saved Excel:\n{xls}")

        elif ev == "error":
            mb.showerror("Error", data)

    def _start(self):
        try:
            import sounddevice as sd

            self._recording = True
            self._frames = []

            self._btn_start.configure(state="disabled")
            self._btn_stop.configure(state="normal")

            def callback(indata, frames, time, status):
                if self._recording:
                    self._frames.append(indata.copy())

                    volume = np.linalg.norm(indata)
                    self._wavedata = np.roll(self._wavedata, -1)
                    self._wavedata[-1] = volume

            self._stream = sd.InputStream(callback=callback)
            self._stream.start()

        except Exception as e:
            mb.showerror("Error", str(e))


    def _stop(self):
        try:
            self._recording = False

            if hasattr(self, "_stream"):
                self._stream.stop()
                self._stream.close()

            import soundfile as sf
            from pathlib import Path
            import uuid
            import numpy as np
            import threading

            Path("temp").mkdir(exist_ok=True)

            wav_path = f"temp/{uuid.uuid4().hex}.wav"
            audio = np.concatenate(self._frames, axis=0)

            sf.write(wav_path, audio, 16000)

            threading.Thread(
                target=self._pipeline,
                args=(wav_path,),
                daemon=True
            ).start()

            self._btn_start.configure(state="normal")
            self._btn_stop.configure(state="disabled")

        except Exception as e:
            mb.showerror("Error", str(e))

    def _load_audio_file(self):
        file_path = fd.askopenfilename()

        if not file_path:
            return

        threading.Thread(
            target=self._pipeline,
            args=(file_path,),
            daemon=True
        ).start()

    def _save_excel(self):
        if not self._session.get("segments"):
            mb.showinfo("Info", "No data to save")
            return

        path = fd.asksaveasfilename(defaultextension=".xlsx")

        if not path:
            return

        export_excel(self._session, path)
        mb.showinfo("Saved", f"Saved to:\n{path}")

if __name__ == "__main__":
    app = SpeakSenseApp()
    app.mainloop()