import os, uuid, threading, datetime
from pathlib import Path
from typing import Dict
import numpy as np
import customtkinter as ctk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
import sounddevice as sd
import soundfile as sf

# ================= UI THEME =================
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

BG = "#F4F7FB"
CARD = "#FFFFFF"
TEXT = "#1E293B"
MUTED = "#64748B"
BLUE = "#2563EB"
RED = "#E53E3E"
GREEN = "#16A34A"
RAISED = "#EDF1F7"

def fmt_time(s):
    m = int(s)//60
    sec = int(s)%60
    return f"{m:02d}:{sec:02d}"

# ================= AUDIO PROCESSOR =================
class AudioProcessor:
    def __init__(self, model_size="base"):
        import whisper
        self.model = whisper.load_model(model_size)

    def process(self, wav_path: str) -> Dict:
        r = self.model.transcribe(str(wav_path))

        segs = [{
            "start": s["start"],
            "end": s["end"],
            "text": s["text"],
            "speaker": 0
        } for s in r["segments"]]

        segs = self.merge_segments(segs)

        return {
            "segments": segs,
            "full_text": " ".join(s["text"] for s in segs)
        }

    def merge_segments(self, segs):
        if not segs:
            return []

        merged = []
        current = segs[0]

        for s in segs[1:]:
            if s["speaker"] == current["speaker"]:
                current["end"] = s["end"]
                current["text"] += " " + s["text"]
            else:
                merged.append(current)
                current = s

        merged.append(current)
        return merged


# ================= SUMMARIZER =================
class Summarizer:
    def summarize(self, text):
        try:
            import ollama
            prompt = f"Summarize this meeting:\n{text[:3000]}"
            r = ollama.chat(model="llama3", messages=[{"role": "user", "content": prompt}])
            return r["message"]["content"]
        except:
            return text[:500]


# ================= EXCEL =================
def export_excel(segments, summary):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Timeline"
    ws["B1"] = "Speaker"
    ws["C1"] = "Text"

    row = 2
    for s in segments:
        ws[f"A{row}"] = f"{fmt_time(s['start'])}-{fmt_time(s['end'])}"
        ws[f"B{row}"] = f"Speaker {s['speaker']+1}"
        ws[f"C{row}"] = s["text"]
        row += 1

    ws[f"A{row+1}"] = "Summary"
    ws[f"B{row+2}"] = summary

    Path("exports").mkdir(exist_ok=True)
    path = f"exports/output_{uuid.uuid4().hex[:6]}.xlsx"
    wb.save(path)

    return path


# ================= APP =================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("SpeakSense")
        self.geometry("900x600")
        self.configure(fg_color=BG)

        self.processor = AudioProcessor()
        self.summarizer = Summarizer()

        self.recording = False
        self.frames = []

        self.build_ui()

    def build_ui(self):
        ctk.CTkLabel(self, text="SpeakSense",
                     font=("Segoe UI", 20, "bold"),
                     text_color=BLUE).pack(pady=10)

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack()

        self.start_btn = ctk.CTkButton(btn_frame, text="Start Recording",
                                      fg_color=RED, command=self.start)
        self.start_btn.pack(side="left", padx=10)

        self.stop_btn = ctk.CTkButton(btn_frame, text="Stop",
                                     command=self.stop, state="disabled")
        self.stop_btn.pack(side="left", padx=10)

        self.load_btn = ctk.CTkButton(btn_frame, text="Load File",
                                     command=self.load_file)
        self.load_btn.pack(side="left", padx=10)

        self.table = ctk.CTkScrollableFrame(self, fg_color=CARD)
        self.table.pack(fill="both", expand=True, padx=10, pady=10)

        self.summary_box = ctk.CTkTextbox(self, height=120)
        self.summary_box.pack(fill="x", padx=10, pady=5)

    # ================= RECORD =================
    def start(self):
        self.recording = True
        self.frames = []

        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")

        def callback(indata, frames, time, status):
            if self.recording:
                self.frames.append(indata.copy())

        self.stream = sd.InputStream(callback=callback)
        self.stream.start()

    def stop(self):
        self.recording = False
        self.stream.stop()

        audio = np.concatenate(self.frames, axis=0)
        Path("temp").mkdir(exist_ok=True)

        path = f"temp/{uuid.uuid4().hex}.wav"
        sf.write(path, audio, 16000)

        self.process_audio(path)

        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")

    # ================= LOAD FILE =================
    def load_file(self):
        file = fd.askopenfilename()
        if file:
            self.process_audio(file)

    # ================= PROCESS =================
    def process_audio(self, path):
        threading.Thread(target=self._run_pipeline, args=(path,), daemon=True).start()

    def _run_pipeline(self, path):
        data = self.processor.process(path)

        self.show_table(data["segments"])

        summary = self.summarizer.summarize(data["full_text"])

        self.summary_box.delete("1.0", "end")
        self.summary_box.insert("end", summary)

        excel = export_excel(data["segments"], summary)

        mb.showinfo("Done", f"Saved Excel:\n{excel}")

    # ================= TABLE =================
    def show_table(self, segments):
        for w in self.table.winfo_children():
            w.destroy()

        headers = ["Timeline", "Speaker", "Text"]

        for i, h in enumerate(headers):
            ctk.CTkLabel(self.table, text=h,
                         font=("Segoe UI", 12, "bold")).grid(row=0, column=i)

        row = 1
        for seg in segments:
            ctk.CTkLabel(self.table,
                         text=f"{fmt_time(seg['start'])}-{fmt_time(seg['end'])}",
                         text_color=MUTED).grid(row=row, column=0)

            ctk.CTkLabel(self.table,
                         text=f"Speaker {seg['speaker']+1}").grid(row=row, column=1)

            ctk.CTkLabel(self.table,
                         text=seg["text"],
                         wraplength=500,
                         justify="left").grid(row=row, column=2)

            row += 1


# ================= RUN =================
if __name__ == "__main__":
    app = App()
    app.mainloop()