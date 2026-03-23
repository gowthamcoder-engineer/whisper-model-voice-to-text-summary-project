import os, re, sys, wave, uuid, time, queue, threading, datetime
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np
import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as fd
import tkinter.messagebox as mb
import customtkinter as ctk

# ── Light / White theme ───────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ─── Palette (WHITE / LIGHT) ──────────────────────────────────────────────────
BG         = "#F4F7FB"   # page background  — very light blue-grey
SURFACE    = "#FFFFFF"   # panels / topbar
CARD       = "#FFFFFF"   # main card background
RAISED     = "#EDF1F7"   # slightly raised surfaces, option-menus
BORDER     = "#CBD5E1"   # dividers, borders
TEXT       = "#1E293B"   # primary text (near-black)
MUTED      = "#64748B"   # secondary / placeholder text
RED        = "#E53E3E"   # record button
GREEN      = "#16A34A"   # done / excel button
BLUE       = "#2563EB"   # logo, progress
PURPLE     = "#7C3AED"
AMBER      = "#D97706"
TEAL       = "#0D9488"
DARK_GREEN = "#14532D"   # excel button text

# ── Speaker accent colours stay vivid; backgrounds & borders go LIGHT ─────────
SPEAKER_FG = ["#0369A1", "#15803D", "#B45309", "#9D174D", "#6D28D9", "#0F766E"]
SPEAKER_BG = ["#EFF6FF", "#F0FDF4", "#FFFBEB", "#FFF0F6", "#F5F3FF", "#F0FDFA"]
SPEAKER_BD = ["#BFDBFE", "#BBF7D0", "#FDE68A", "#FBCFE8", "#DDD6FE", "#99F6E4"]

def spk_fg(i): return SPEAKER_FG[i % len(SPEAKER_FG)]
def spk_bg(i): return SPEAKER_BG[i % len(SPEAKER_BG)]
def spk_bd(i): return SPEAKER_BD[i % len(SPEAKER_BD)]
def spk_label(i): return f"Speaker {i+1}"
def fmt_time(s):
    m = int(s)//60; sec = int(s)%60
    return f"{m:02d}:{sec:02d}"

# ─── Audio Processor ──────────────────────────────────────────────────────────
class AudioProcessor:
    def __init__(self, model_size="base"):
        import whisper
        self.model = whisper.load_model(model_size)
        self.model_size = model_size

    def process(self, wav_path: str) -> Dict:
        r = self.model.transcribe(str(wav_path), word_timestamps=True, verbose=False)
        segs = [{"id": s["id"], "start": round(s["start"],2),
                 "end": round(s["end"],2), "text": s["text"].strip(),
                 "speaker": 0}
                for s in r["segments"]]
        segs = self._diarize(segs)
        spk_data = self._build_speakers(segs)
        return {
            "segments":     segs,
            "speaker_data": spk_data,
            "full_text":    " ".join(s["text"] for s in segs),
            "duration":     segs[-1]["end"] if segs else 0,
        }

    def _diarize(self, segs):
        idx = 0
        for i, s in enumerate(segs):
            if i > 0 and (s["start"] - segs[i-1]["end"]) >= 1.2:
                idx = (idx + 1) % 4
            s["speaker"] = idx
        return segs

    def _build_speakers(self, segs) -> Dict[int, Dict]:
        data = {}
        for s in segs:
            idx = s["speaker"]
            if idx not in data:
                data[idx] = {"text":"","segments":[],"duration":0.0,"count":0,
                             "first":s["start"],"last":s["end"]}
            d = data[idx]
            d["text"] += (" " + s["text"]).lstrip()
            d["segments"].append({"start":s["start"],"end":s["end"],"text":s["text"]})
            d["duration"] += round(s["end"] - s["start"], 2)
            d["count"] += 1
            d["last"] = max(d["last"], s["end"])
        for d in data.values(): d["text"] = d["text"].strip()
        return dict(sorted(data.items()))

# ─── Summarizer ───────────────────────────────────────────────────────────────
class Summarizer:
    STOP = {"the","a","an","in","on","at","to","for","of","and","or","but","is","are",
            "was","were","be","been","it","this","that","with","i","we","you","he",
            "she","they","so","as","by","from","not","have","has","had","do","did"}

    def ollama_sum(self, text, model, max_w=160, hint=""):
        import ollama
        p = (f"Summarise in {max_w} words or fewer. Concise, accurate. {hint}\n\n"
             f"TEXT:\n{text[:4000]}\n\nSUMMARY:")
        r = ollama.chat(model=model, messages=[{"role":"user","content":p}])
        return r["message"]["content"].strip()

    def bart_sum(self, text):
        try:
            from transformers import pipeline
            p = pipeline("summarization","sshleifer/distilbart-cnn-12-6",device=-1)
            return p(text[:3500],max_length=220,min_length=60,do_sample=False)[0]["summary_text"].strip()
        except:
            return self._extractive(text, 4)

    def _extractive(self, text, n=4):
        sents = [s.strip() for s in re.split(r"(?<=[.!?])\s+",text) if len(s.split())>=5]
        if not sents: return text[:400]
        freq = Counter(re.findall(r"\b\w+\b",text.lower()))
        scored = []
        for s in sents:
            words = [w for w in re.findall(r"\b\w+\b",s.lower()) if w not in self.STOP]
            scored.append((sum(freq[w] for w in words)/(len(words)+1), s))
        scored.sort(reverse=True)
        top = sorted(scored[:n], key=lambda x: sents.index(x[1]))
        return " ".join(x[1] for x in top)

    def speaker_sum(self, text, idx, model):
        if len(text.split()) < 15: return text.strip()
        try:
            return self.ollama_sum(text, model, 80,
                                   f"Focus only on what {spk_label(idx)} said.")
        except:
            return self._extractive(text, 2)

    def best_overall(self, full_text, spk_summaries: Dict[int,str], model) -> str:
        combined = "\n".join(f"[{spk_label(i)}]: {s}" for i,s in spk_summaries.items())
        try:
            return self.ollama_sum(
                combined, model, 200,
                "Extract the most important points from ALL speakers. "
                "Write a unified meeting/conversation summary covering key decisions, "
                "topics, and agreements. Do NOT attribute to speakers—write as a whole."
            )
        except:
            return self._extractive(full_text, 6)

# ─── Excel Export ─────────────────────────────────────────────────────────────
def export_excel(session: Dict, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Session Results"

    def fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))
    def bdr():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s,right=s,top=s,bottom=s)

    spk_data      = session["speaker_data"]
    spk_summaries = session["speaker_summaries"]
    segments      = session["segments"]
    overall       = session["overall_summary"]
    n_spk         = len(spk_data)

    ws.column_dimensions["A"].width = 14
    for i in range(n_spk):
        ws.column_dimensions[get_column_letter(2+i)].width = 52

    last_col = get_column_letter(1 + n_spk)
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value     = f"SpeakSense Recording  —  {datetime.datetime.now():%Y-%m-%d %H:%M}"
    c.font      = Font(name="Calibri", bold=True, size=15, color="FFFFFF")
    c.fill      = fill("#0D1B2A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    SPK_HDR_BG = ["1A5276","1E8449","784212","76448A","117A65","1A5276"]
    c = ws.cell(2,1,"TIMELINE")
    c.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    c.fill=fill("1B2631"); c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=bdr()
    ws.row_dimensions[2].height = 26
    for i in range(n_spk):
        c = ws.cell(2, 2+i, f"SPEAKER {i+1}")
        c.font = Font(name="Calibri",bold=True,size=11,color="FFFFFF")
        c.fill = fill(SPK_HDR_BG[i % len(SPK_HDR_BG)])
        c.alignment = Alignment(horizontal="center",vertical="center")
        c.border = bdr()

    SPK_LIGHT = ["D6EAF8","D5F5E3","FDEBD0","F5EEF8","D1F2EB","D6EAF8"]
    SPK_DARK  = ["1A5276","1E8449","784212","76448A","117A65","1A5276"]

    row = 3
    for seg in segments:
        idx  = seg["speaker"]
        col  = 2 + idx
        light = SPK_LIGHT[idx % len(SPK_LIGHT)]
        dark  = SPK_DARK [idx % len(SPK_DARK)]

        c = ws.cell(row,1, f"{fmt_time(seg['start'])}→{fmt_time(seg['end'])}")
        c.font=Font(name="Courier New",size=9,color="666666")
        c.fill=fill("F2F3F4"); c.alignment=Alignment(horizontal="center",vertical="top")
        c.border=bdr(); ws.row_dimensions[row].height=40

        c = ws.cell(row, col, seg["text"])
        c.font=Font(name="Calibri",size=10)
        c.fill=fill(light); c.alignment=Alignment(vertical="top",wrap_text=True)
        c.border=bdr()

        for j in range(n_spk):
            if j != idx:
                c = ws.cell(row,2+j,"")
                c.fill=fill("F8F9FA"); c.border=bdr()
        row += 1

    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row,1,"  📝  SPEAKER SUMMARIES")
    c.font=Font(name="Calibri",bold=True,size=11,color="FFFFFF")
    c.fill=fill("1B2631"); c.alignment=Alignment(vertical="center")
    c.border=bdr(); ws.row_dimensions[row].height=24
    row += 1

    ws.cell(row,1,"SUMMARY").font=Font(name="Calibri",bold=True,size=9,color="AAAAAA")
    ws.cell(row,1).fill=fill("F2F3F4"); ws.cell(row,1).border=bdr()
    ws.cell(row,1).alignment=Alignment(horizontal="center",vertical="top")
    for i, (idx, summ) in enumerate(spk_summaries.items()):
        c = ws.cell(row, 2+i, summ)
        c.font=Font(name="Calibri",size=10,italic=True)
        c.fill=fill(SPK_LIGHT[idx%len(SPK_LIGHT)])
        c.alignment=Alignment(vertical="top",wrap_text=True)
        c.border=bdr()
    ws.row_dimensions[row].height = 80
    row += 1

    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row,1,"  ✅  OVERALL BEST SUMMARY — Key Points from All Speakers")
    c.font=Font(name="Calibri",bold=True,size=12,color="FFFFFF")
    c.fill=fill("145A32"); c.alignment=Alignment(vertical="center")
    c.border=bdr(); ws.row_dimensions[row].height=26
    row += 1

    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row,1, overall)
    c.font=Font(name="Calibri",size=11)
    c.fill=fill("EAFAF1"); c.alignment=Alignment(vertical="top",wrap_text=True)
    c.border=bdr()
    ws.row_dimensions[row].height = max(80, min(160, len(overall)//3+20))

    ws.freeze_panes = "A3"
    wb.save(path)
    print(f"[Excel] Saved → {path}")


# ─── Main App ─────────────────────────────────────────────────────────────────
class SpeakSenseApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("SpeakSense")
        self.geometry("1400x860")
        self.minsize(1000, 640)
        self.configure(fg_color=BG)
        self.resizable(True, True)

        self._recording   = False
        self._frames      = []
        self._timer_sec   = 0
        self._timer_job   = None
        self._wavedata    = np.zeros(300)
        self._processor   = None
        self._summarizer  = Summarizer()
        self._session     = {}
        self._row_widgets = []
        self._spk_cols    = []
        self._n_speakers  = 0

        self._build_topbar()
        self._build_controls()
        self._build_waveform()
        self._build_main_table()
        self._waveform_loop()

    # ── Top bar ───────────────────────────────────────────────────────────
    def _build_topbar(self):
        bar = ctk.CTkFrame(self, fg_color=SURFACE, height=52, corner_radius=0,
                           border_width=1, border_color=BORDER)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        ctk.CTkLabel(bar, text="🎙  SpeakSense",
                     font=ctk.CTkFont("Segoe UI",19,"bold"),
                     text_color=BLUE).pack(side="left", padx=22, pady=12)

        right = ctk.CTkFrame(bar, fg_color="transparent")
        right.pack(side="right", padx=18)

        self._timer_lbl = ctk.CTkLabel(right, text="00:00",
                                        font=ctk.CTkFont("Courier",22,"bold"),
                                        text_color=TEXT)
        self._timer_lbl.pack(side="right", padx=(12,0))

        self._status_lbl = ctk.CTkLabel(right, text="● READY",
                                         font=ctk.CTkFont("Courier",11,"bold"),
                                         text_color=MUTED)
        self._status_lbl.pack(side="right")

    # ── Controls row ──────────────────────────────────────────────────────
    def _build_controls(self):
        bar = ctk.CTkFrame(self, fg_color=RAISED, height=62, corner_radius=0,
                           border_width=1, border_color=BORDER)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        ctk.CTkLabel(bar, text="WHISPER",
                     font=ctk.CTkFont("Courier",9), text_color=MUTED
                     ).pack(side="left", padx=(18,3))
        self._w_var = ctk.StringVar(value="base")
        ctk.CTkOptionMenu(bar, variable=self._w_var,
                          values=["tiny","base","small","medium","large"],
                          width=120, height=30,
                          fg_color=SURFACE, button_color=BORDER,
                          text_color=TEXT, button_hover_color=BORDER,
                          dropdown_fg_color=SURFACE, dropdown_text_color=TEXT,
                          font=ctk.CTkFont("Courier",11)
                          ).pack(side="left", padx=(0,18))

        ctk.CTkLabel(bar, text="OLLAMA",
                     font=ctk.CTkFont("Courier",9), text_color=MUTED
                     ).pack(side="left", padx=(0,3))
        self._o_var = ctk.StringVar(value="llama3")
        ctk.CTkOptionMenu(bar, variable=self._o_var,
                          values=["llama3","llama3.1","mistral","gemma2","phi3","qwen2"],
                          width=130, height=30,
                          fg_color=SURFACE, button_color=BORDER,
                          text_color=TEXT, button_hover_color=BORDER,
                          dropdown_fg_color=SURFACE, dropdown_text_color=TEXT,
                          font=ctk.CTkFont("Courier",11)
                          ).pack(side="left", padx=(0,28))

        self._btn_start = ctk.CTkButton(
            bar, text="⏺  Start Recording", width=170, height=34,
            fg_color=RED, hover_color="#b91c1c",
            font=ctk.CTkFont("Segoe UI",13,"bold"),
            command=self._start)
        self._btn_start.pack(side="left", padx=4)

        self._btn_stop = ctk.CTkButton(
            bar, text="⏹  Stop & Analyze", width=170, height=34,
            state="disabled", fg_color=BORDER, hover_color="#94a3b8",
            text_color=TEXT,
            font=ctk.CTkFont("Segoe UI",13,"bold"),
            command=self._stop)
        self._btn_stop.pack(side="left", padx=4)

        self._btn_xl = ctk.CTkButton(
            bar, text="📊  Save Excel", width=150, height=34,
            state="disabled", fg_color=GREEN, hover_color="#15803d",
            text_color="#FFFFFF",
            font=ctk.CTkFont("Segoe UI",13,"bold"),
            command=self._save_excel)
        self._btn_xl.pack(side="left", padx=4)

        # Progress (hidden)
        self._prog_frame = ctk.CTkFrame(bar, fg_color="transparent")
        self._prog_lbl = ctk.CTkLabel(self._prog_frame, text="",
                                       font=ctk.CTkFont("Courier",10),
                                       text_color=BLUE)
        self._prog_lbl.pack(side="left", padx=(16,6))
        self._prog_bar = ctk.CTkProgressBar(self._prog_frame, width=200, height=10,
                                             fg_color=BORDER, progress_color=BLUE)
        self._prog_bar.pack(side="left")
        self._prog_bar.set(0)

    # ── Waveform ──────────────────────────────────────────────────────────
    def _build_waveform(self):
        self._wave_frame = ctk.CTkFrame(self, fg_color=SURFACE, height=58,
                                         corner_radius=0,
                                         border_width=1, border_color=BORDER)
        self._wave_frame.pack(fill="x")
        self._wave_frame.pack_propagate(False)
        # Canvas uses SURFACE as bg so it blends with light theme
        self._wave_canvas = tk.Canvas(self._wave_frame, bg=SURFACE,
                                       highlightthickness=0, height=58)
        self._wave_canvas.pack(fill="both", expand=True, padx=16, pady=4)

    def _waveform_loop(self):
        try:
            c = self._wave_canvas
            W = c.winfo_width() or 900; H = 58
            c.delete("all")
            if self._recording:
                n = len(self._wavedata); bw = W/n
                for i, v in enumerate(self._wavedata):
                    bh = max(2, v * H * 3.5)
                    x  = i * bw
                    # On white bg: use solid red bars that darken with amplitude
                    intensity = int(200 + v * 55)          # 200-255 range for R
                    blue_ch   = int(max(0, 80 - v * 80))   # fade blue out
                    color = f"#{intensity:02x}{blue_ch:02x}{blue_ch:02x}"
                    c.create_rectangle(x, H/2-bh/2, x+bw-1, H/2+bh/2,
                                       fill=color, outline="")
            else:
                c.create_text(W/2, H/2,
                              text="── waveform appears during recording ──",
                              fill=MUTED, font=("Segoe UI",10))
        except Exception: pass
        self.after(45, self._waveform_loop)

    # ── Main table ────────────────────────────────────────────────────────
    def _build_main_table(self):
        outer = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        outer.pack(fill="both", expand=True, padx=14, pady=(10,10))

        self._scroll = ctk.CTkScrollableFrame(
            outer, fg_color=CARD, corner_radius=12,
            border_width=1, border_color=BORDER,
            scrollbar_button_color=BORDER,
            scrollbar_button_hover_color="#94a3b8")
        self._scroll.pack(fill="both", expand=True)

        self._placeholder = ctk.CTkLabel(
            self._scroll,
            text="Press  ⏺ Start Recording  to begin.\n\n"
                 "Columns will appear automatically:\n"
                 "Timeline  |  Speaker 1  |  Speaker 2  |  …  |  Speaker N\n\n"
                 "Each speaker's spoken text appears live.\n"
                 "Summaries and the overall best summary are shown after analysis.\n"
                 "Excel is saved automatically when ready.",
            font=ctk.CTkFont("Segoe UI",14),
            text_color=MUTED, justify="center")
        self._placeholder.pack(expand=True, pady=80)

    def _rebuild_table(self):
        for w in self._scroll.winfo_children():
            w.destroy()

        spk_data = self._session.get("speaker_data", {})
        self._n_speakers = len(spk_data)
        if self._n_speakers == 0: return

        grid = self._scroll
        grid.grid_columnconfigure(0, weight=0, minsize=100)
        for i in range(self._n_speakers):
            grid.grid_columnconfigure(i+1, weight=1, minsize=260)

        # Header row
        tl_hdr = ctk.CTkLabel(grid, text="TIMELINE",
                               font=ctk.CTkFont("Courier",9,"bold"),
                               text_color=MUTED,
                               fg_color=RAISED, corner_radius=6, width=90)
        tl_hdr.grid(row=0, column=0, padx=(4,2), pady=(4,2), sticky="nsew")

        self._spk_header_labels = {}
        for i in range(self._n_speakers):
            fg = spk_fg(i); bg = spk_bg(i); bd = spk_bd(i)
            hdr = ctk.CTkFrame(grid, fg_color=bg, corner_radius=8,
                               border_width=1, border_color=bd)
            hdr.grid(row=0, column=i+1, padx=2, pady=(4,2), sticky="nsew")
            ctk.CTkLabel(hdr, text=f"  SPEAKER {i+1}",
                         font=ctk.CTkFont("Segoe UI",13,"bold"),
                         text_color=fg).pack(side="left", padx=10, pady=8)
            info = spk_data.get(i, {})
            ctk.CTkLabel(hdr,
                         text=f"{info.get('count',0)} segs · {info.get('duration',0):.1f}s",
                         font=ctk.CTkFont("Courier",9), text_color=MUTED
                         ).pack(side="left", padx=4, pady=8)

        self._seg_row_start = 1
        self._next_seg_row  = 1
        self._seg_cells     = {}

    def _add_segment_row(self, seg: Dict):
        grid = self._scroll
        row  = self._next_seg_row
        self._next_seg_row += 1
        idx  = seg["speaker"]
        col  = idx + 1

        # Timeline cell
        tl = ctk.CTkLabel(grid,
                           text=f"{fmt_time(seg['start'])}\n→{fmt_time(seg['end'])}",
                           font=ctk.CTkFont("Courier",10),
                           text_color=MUTED, fg_color=RAISED,
                           corner_radius=4, width=90, height=52)
        tl.grid(row=row, column=0, padx=(4,2), pady=1, sticky="nsew")

        fg = spk_fg(idx); bg = spk_bg(idx); bd = spk_bd(idx)
        cell = ctk.CTkFrame(grid, fg_color=bg, corner_radius=6,
                             border_width=1, border_color=bd)
        cell.grid(row=row, column=col, padx=2, pady=1, sticky="nsew")
        ctk.CTkLabel(cell, text=seg["text"],
                     font=ctk.CTkFont("Segoe UI",11),
                     text_color=TEXT, wraplength=340, justify="left",
                     anchor="nw").pack(anchor="nw", padx=10, pady=7, fill="both")

        # Empty cells for other speakers
        for j in range(self._n_speakers):
            if j != idx:
                e = ctk.CTkFrame(grid, fg_color="#F8FAFC", corner_radius=4, height=52)
                e.grid(row=row, column=j+1, padx=2, pady=1, sticky="nsew")

    def _build_summary_section(self):
        grid = self._scroll
        row  = self._next_seg_row

        # Divider
        sep = ctk.CTkFrame(grid, fg_color=BORDER, height=2, corner_radius=0)
        sep.grid(row=row, column=0, columnspan=self._n_speakers+1,
                 padx=4, pady=(8,4), sticky="ew")
        row += 1

        # "SPEAKER SUMMARIES" label
        lbl = ctk.CTkLabel(grid, text="📝  SPEAKER SUMMARIES",
                            font=ctk.CTkFont("Segoe UI",12,"bold"),
                            text_color=AMBER, fg_color=RAISED,
                            corner_radius=6, height=32, anchor="w")
        lbl.grid(row=row, column=0, columnspan=self._n_speakers+1,
                 padx=4, pady=(0,4), sticky="ew")
        row += 1

        # Timeline column label for summary row
        ctk.CTkLabel(grid, text="SUMMARY", font=ctk.CTkFont("Courier",8),
                     text_color=MUTED, fg_color=RAISED, corner_radius=4,
                     width=90).grid(row=row, column=0, padx=(4,2), pady=2, sticky="nsew")

        self._spk_sum_labels = {}
        for i in range(self._n_speakers):
            fg = spk_fg(i); bg = spk_bg(i); bd = spk_bd(i)
            frame = ctk.CTkFrame(grid, fg_color=bg, corner_radius=8,
                                  border_width=1, border_color=bd)
            frame.grid(row=row, column=i+1, padx=2, pady=2, sticky="nsew")
            ctk.CTkLabel(frame, text=f"Speaker {i+1} Summary",
                         font=ctk.CTkFont("Courier",9,"bold"),
                         text_color=fg).pack(anchor="w", padx=10, pady=(8,2))
            lbl2 = ctk.CTkLabel(frame, text="Generating…",
                                 font=ctk.CTkFont("Segoe UI",11),
                                 text_color=TEXT, wraplength=340,
                                 justify="left", anchor="nw")
            lbl2.pack(anchor="nw", padx=10, pady=(0,10), fill="both")
            self._spk_sum_labels[i] = lbl2
        row += 1

        # Overall divider
        sep2 = ctk.CTkFrame(grid, fg_color=BORDER, height=2, corner_radius=0)
        sep2.grid(row=row, column=0, columnspan=self._n_speakers+1,
                  padx=4, pady=(10,4), sticky="ew")
        row += 1

        # Overall header — light green on white background
        hdr_overall = ctk.CTkLabel(
            grid,
            text="✅  OVERALL BEST SUMMARY — Key Points from All Speakers",
            font=ctk.CTkFont("Segoe UI",13,"bold"),
            text_color=GREEN, fg_color="#DCFCE7",      # light green bg
            corner_radius=8, height=36, anchor="w")
        hdr_overall.grid(row=row, column=0, columnspan=self._n_speakers+1,
                         padx=4, pady=(0,2), sticky="ew")
        row += 1

        # Overall summary text box — light green fill, dark text
        overall_frame = ctk.CTkFrame(grid, fg_color="#F0FDF4", corner_radius=10,
                                      border_width=1, border_color="#86EFAC")
        overall_frame.grid(row=row, column=0, columnspan=self._n_speakers+1,
                           padx=4, pady=(0,12), sticky="ew")
        self._overall_lbl = ctk.CTkLabel(
            overall_frame, text="Generating overall summary…",
            font=ctk.CTkFont("Segoe UI",12),
            text_color=DARK_GREEN,              # dark green text, readable on white
            wraplength=1100,
            justify="left", anchor="nw")
        self._overall_lbl.pack(anchor="nw", padx=16, pady=14, fill="both")

    # ── Recording ─────────────────────────────────────────────────────────
    def _start(self):
        import sounddevice as sd
        self._frames.clear()
        self._recording = True
        self._session   = {}
        self._btn_start.configure(state="disabled", fg_color="#b91c1c")
        self._btn_stop.configure(state="normal")
        self._btn_xl.configure(state="disabled")
        self._status("⏺  RECORDING", RED)
        self._timer_sec = 0; self._tick()

        self._stream = sd.InputStream(
            samplerate=16000, channels=1, dtype="int16",
            blocksize=4096, callback=self._audio_cb)
        self._stream.start()

    def _audio_cb(self, indata, frames, time_info, status):
        if self._recording:
            self._frames.append(indata.copy())
            v = np.abs(indata[:,0].astype(np.float32)/32768).mean()
            self._wavedata = np.roll(self._wavedata, -1)
            self._wavedata[-1] = v

    def _stop(self):
        if not self._recording: return
        self._recording = False
        self._btn_stop.configure(state="disabled")
        if hasattr(self, "_stream"): self._stream.stop(); self._stream.close()
        if self._timer_job: self.after_cancel(self._timer_job)

        if not self._frames:
            mb.showerror("No Audio","No audio captured — check your microphone.")
            self._btn_start.configure(state="normal", fg_color=RED)
            return

        Path("recordings").mkdir(exist_ok=True)
        wav_path = f"recordings/{uuid.uuid4().hex[:8]}.wav"
        audio = np.concatenate(self._frames, axis=0)
        with wave.open(wav_path,"wb") as wf:
            wf.setnchannels(1); wf.setsampwidth(2)
            wf.setframerate(16000); wf.writeframes(audio.tobytes())

        self._status("⚙  PROCESSING", AMBER)
        self._show_progress(True, 0.05, "Loading Whisper…")
        threading.Thread(target=self._pipeline, args=(wav_path,), daemon=True).start()

    # ── Pipeline ──────────────────────────────────────────────────────────
    def _pipeline(self, wav_path):
        try:
            wm = self._w_var.get()
            om = self._o_var.get()

            self._emit("progress",(0.15,"🎙 Transcribing with Whisper…"))
            if self._processor is None or self._processor.model_size != wm:
                self._processor = AudioProcessor(wm)

            self._emit("progress",(0.30,"🗣 Identifying speakers…"))
            data = self._processor.process(wav_path)
            self._session.update(data)
            self._emit("transcript_ready", data)

            self._emit("progress",(0.52,"📝 Summarizing speakers…"))
            spk_summaries = {}
            for i, info in data["speaker_data"].items():
                s = self._summarizer.speaker_sum(info["text"], i, om)
                spk_summaries[i] = s
                self._emit("spk_summary",(i, s))

            self._emit("progress",(0.78,"✨ Generating overall best summary…"))
            overall = self._summarizer.best_overall(data["full_text"], spk_summaries, om)

            self._session["speaker_summaries"] = spk_summaries
            self._session["overall_summary"]   = overall

            self._emit("progress",(0.95,"💾 Saving Excel…"))
            ts  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            Path("exports").mkdir(exist_ok=True)
            xls = f"exports/SpeakSense_{ts}.xlsx"
            export_excel(self._session, xls)

            self._emit("complete", (overall, xls))

        except Exception as e:
            import traceback; traceback.print_exc()
            self._emit("error", str(e))

    # ── Event dispatch ────────────────────────────────────────────────────
    def _emit(self, ev, data):
        self.after(0, lambda: self._handle(ev, data))

    def _handle(self, ev, data):
        if ev == "progress":
            pct, label = data
            self._prog_bar.set(pct)
            self._prog_lbl.configure(text=label)

        elif ev == "transcript_ready":
            self._rebuild_table()
            for seg in data["segments"]:
                self._add_segment_row(seg)
            self._build_summary_section()

        elif ev == "spk_summary":
            i, summ = data
            if hasattr(self,"_spk_sum_labels") and i in self._spk_sum_labels:
                self._spk_sum_labels[i].configure(text=summ)

        elif ev == "complete":
            overall, xls = data
            if hasattr(self,"_overall_lbl"):
                self._overall_lbl.configure(text=overall)
            self._show_progress(False)
            self._btn_start.configure(state="normal", fg_color=RED)
            self._btn_xl.configure(state="normal")
            self._status("✅  DONE", GREEN)
            mb.showinfo("Done! 🎉",
                        f"Analysis complete!\nExcel saved automatically to:\n\n{xls}\n\n"
                        "You can also click 📊 Save Excel to save again.")

        elif ev == "error":
            self._show_progress(False)
            self._btn_start.configure(state="normal", fg_color=RED)
            self._status("✗  ERROR", RED)
            mb.showerror("Error", str(data))

    # ── Helpers ───────────────────────────────────────────────────────────
    def _status(self, text, color):
        self._status_lbl.configure(text=text, text_color=color)

    def _show_progress(self, show, pct=0, label=""):
        if show:
            self._prog_frame.pack(side="left", padx=(10,0))
            self._prog_bar.set(pct)
            self._prog_lbl.configure(text=label)
        else:
            self._prog_frame.pack_forget()

    def _tick(self):
        if self._recording:
            self._timer_sec += 1
            self._timer_lbl.configure(
                text=f"{self._timer_sec//60:02d}:{self._timer_sec%60:02d}")
            self._timer_job = self.after(1000, self._tick)

    def _save_excel(self):
        if not self._session.get("segments"):
            mb.showinfo("Nothing to save","Process a recording first.")
            return
        path = fd.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"SpeakSense_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx",
            title="Save Excel")
        if not path: return
        try:
            export_excel(self._session, path)
            mb.showinfo("Saved!", f"Saved to:\n{path}")
        except Exception as e:
            mb.showerror("Failed", str(e))


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SpeakSenseApp()
    app.mainloop()
